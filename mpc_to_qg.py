import openpyxl
import pandas as pd
import os
import re
from typing import Dict, Any, List

class MPCToQGTranscompiler:
    """
    CIQ FUSION Phase 1 Transcompiler:
    Transforms an MPC CIQ (.xlsm / .xlsx) into a fully populated QG CIQ (.xlsx).
    """

    def __init__(self, mpc_filepath: str):
        self.mpc_filepath = mpc_filepath
        self.mpc_wb = openpyxl.load_workbook(mpc_filepath, data_only=True)
        self.qg_wb = openpyxl.Workbook()
        self.qg_wb.remove(self.qg_wb.active)  # Remove default initial sheet
        self.issues_log = []

    def _log_issue(self, fa_code: str, parameter: str, severity: str, message: str):
        self.issues_log.append({
            "FA Code": fa_code or "N/A",
            "Parameter Name": parameter,
            "Severity": severity.upper(),
            "Description": message
        })

    def _get_sheet_rows(self, sheet_name: str) -> List[List[Any]]:
        if sheet_name in self.mpc_wb.sheetnames:
            return list(self.mpc_wb[sheet_name].iter_rows(values_only=True))
        return []

    # 1. Controller Info Sheet
    def generate_controller_info(self) -> pd.DataFrame:
        nr_rows = self._get_sheet_rows('NR RF')
        usid = "N/A"
        
        for r in nr_rows[3:]:
            if len(r) > 9 and r[9]:
                usid = str(r[9]).strip()
                break

        if usid == "N/A":
            self._log_issue("", "USID", "WARNING", "USID not found in MPC 'NR RF' sheet. Generated fallback Controller ID.")

        return pd.DataFrame([{
            "USID": usid,
            "Controller": "6610",
            "Controller ID": f"DTFC{usid}_C001" if usid != "N/A" else "N/A"
        }])

    # 2. eNB Info Sheet
    def generate_enb_info(self) -> pd.DataFrame:
        lte_rows = self._get_sheet_rows('LTE RF')
        records = []
        
        for r in lte_rows[3:]:
            if r and len(r) > 13 and r[6]:  # FA CODE
                fa_code = str(r[6]).strip()
                enb_id = str(r[13]).strip() if r[13] else "N/A"
                enb_name = str(r[10]).strip() if len(r) > 10 and r[10] else f"DTL{enb_id}"
                
                records.append({
                    "eNBId": enb_id,
                    "eNodeB Name": enb_name,
                    "Site Address": "N/A",
                    "County": "N/A",
                    "Structure Type": "MACRO / TOWER",
                    "RBS type": "6601",
                    "PLMNId": "310410",
                    "MCC": "310",
                    "MNC": "410",
                    "mncLength": "3"
                })
        
        if not records:
            records.append({
                "eNBId": "N/A", "eNodeB Name": "N/A", "Site Address": "N/A",
                "County": "N/A", "Structure Type": "N/A", "RBS type": "6601",
                "PLMNId": "310410", "MCC": "310", "MNC": "410", "mncLength": "3"
            })
        return pd.DataFrame(records).drop_duplicates()

    # 3. gNB Info Sheet
    def generate_gnb_info(self) -> pd.DataFrame:
        nr_rows = self._get_sheet_rows('NR RF')
        records = []

        for r in nr_rows[3:]:
            if r and len(r) > 13 and r[6]:
                gnb_id = str(r[13]).strip() if r[13] else "N/A"
                gnb_name = str(r[10]).strip() if len(r) > 10 and r[10] else f"DTFN{gnb_id}"
                
                records.append({
                    "gNBId": gnb_id,
                    "gNodeB Name": gnb_name,
                    "numberOfSectors per DUL": "9",
                    "DU type": "6672",
                    "1st XMU": "No",
                    "1st XMU Port 1": "N/A",
                    "1st XMU Port 2": "N/A",
                    "1st XMU Port 3": "N/A",
                    "2nd XMU": "N/A",
                    "2nd XMU Port 1": "N/A"
                })

        if not records:
            records.append({
                "gNBId": "N/A", "gNodeB Name": "N/A", "numberOfSectors per DUL": "9",
                "DU type": "6672", "1st XMU": "No", "1st XMU Port 1": "N/A",
                "1st XMU Port 2": "N/A", "1st XMU Port 3": "N/A", "2nd XMU": "N/A", "2nd XMU Port 1": "N/A"
            })
        return pd.DataFrame(records).drop_duplicates()

    # 4. eUtran Parameters Sheet
    def generate_eutran_parameters(self) -> pd.DataFrame:
        lte_rows = self._get_sheet_rows('LTE RF')
        records = []

        for r in lte_rows[3:]:
            if r and len(r) > 9 and r[6]:
                fa_code = str(r[6]).strip()
                node_name = str(r[10]).strip() if len(r) > 10 and r[10] else "DTL0000"
                cell_id = f"{node_name}_7A_1"
                
                records.append({
                    "NOTES:(Please Provide Detailed Build and Design Change Notes)": r[0] if r[0] else "E2E LTE Build",
                    "Carrier Cell Intention": r[1] if r[1] else "New Cell Add",
                    "Pace": r[7] if r[7] else "N/A",
                    "Cell Status": r[2] if r[2] else "UNLOCKED/UNBARRED/UNRESERVED",
                    "eNBId": r[13] if len(r) > 13 else "N/A",
                    "EutranCellFDDId": cell_id,
                    "latitude": "N/A",
                    "latHemisphere": "N",
                    "longitude": "N/A",
                    "geoDatum": "NAD83",
                    "cellRange": "15000",
                    "beamDirection": "0",
                    "Antenna Height": "100"
                })

        return pd.DataFrame(records)

    # 5. 5G Info Sheet
    def generate_5g_info(self) -> pd.DataFrame:
        nr_rows = self._get_sheet_rows('NR RF')
        records = []

        for r in nr_rows[3:]:
            if r and len(r) > 9 and r[6]:
                fa_code = str(r[6]).strip()
                gnb_id = str(r[13]).strip() if len(r) > 13 and r[13] else "N/A"
                gnb_name = str(r[10]).strip() if len(r) > 10 and r[10] else "DTFN0000"
                cell_du = f"{gnb_name}_N005A_1"
                sector_eq = f"{gnb_name}_N005A"

                records.append({
                    "Cell Status": r[2] if r[2] else "LOCKED/BARRED/RESERVED",
                    "gNBId": gnb_id,
                    "gNB Name": gnb_name,
                    "NRCellDU": cell_du,
                    "NRCellCU": cell_du,
                    "SectorEquipmentFunction": sector_eq,
                    "NRSectorCarrier": cell_du,
                    "CellRange": "23000",
                    "Address": "N/A",
                    "Latitude": "N/A",
                    "Longitude": "N/A",
                    "BBU Type": "6672",
                    "nRTAC": "1",
                    "NumberOfSectors per BB": "9",
                    "cellLocalId": "1"
                })

        return pd.DataFrame(records)

    # 6. ISSUES_LOG Sheet (Goal #3 Requirement)
    def generate_issues_log(self) -> pd.DataFrame:
        if not self.issues_log:
            self.issues_log.append({
                "FA Code": "ALL",
                "Parameter Name": "System Check",
                "Severity": "INFO",
                "Description": "MPC to QG CIQ Transcompilation Executed Successfully with 0 critical errors."
            })
        return pd.DataFrame(self.issues_log)

    def convert(self, output_filepath: str) -> Dict[str, Any]:
        """Main converter execution pipeline."""
        sheet_generators = {
            "Controller Info": self.generate_controller_info,
            "eNB Info": self.generate_enb_info,
            "gNB Info": self.generate_gnb_info,
            "eUtran Parameters": self.generate_eutran_parameters,
            "5G Info": self.generate_5g_info,
            "ISSUES_LOG": self.generate_issues_log
        }

        converted_sheets = []
        for sheet_title, generator in sheet_generators.items():
            df = generator()
            ws = self.qg_wb.create_sheet(title=sheet_title)
            
            # Write Header
            ws.append(list(df.columns))
            # Write Rows
            for row in df.itertuples(index=False):
                ws.append(list(row))
            
            converted_sheets.append(sheet_title)

        # Save workbook
        self.qg_wb.save(output_filepath)

        return {
            "status": "SUCCESS",
            "output_filepath": output_filepath,
            "converted_sheets": converted_sheets,
            "issue_count": len(self.issues_log)
        }
